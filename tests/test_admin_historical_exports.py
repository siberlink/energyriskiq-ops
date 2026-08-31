import asyncio
import io
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
from fastapi import HTTPException

from src.api import admin_historical_exports as exports
from src.api import admin_routes


async def _response_bytes(response):
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk.encode() if isinstance(chunk, str) else chunk)
    return b"".join(chunks)


def test_catalog_contains_requested_and_supporting_datasets():
    keys = {item["key"] for item in exports.export_catalog()}
    assert {
        "geri-live",
        "geri-daily",
        "brent-daily",
        "brent-intraday",
        "vix-daily",
        "lng-daily",
        "ttf-daily",
        "wti-daily",
        "wti-intraday",
        "natgas-intraday",
        "eu-storage",
        "country-storage",
        "geri-drivers",
        "alert-events",
        "eeri-daily",
        "egsi-m-daily",
        "egsi-s-daily",
    } <= keys


def test_unknown_dataset_and_format_fail_without_querying(monkeypatch):
    monkeypatch.setattr(
        exports,
        "_fetch_rows",
        lambda *_args: pytest.fail("database should not be queried"),
    )
    with pytest.raises(HTTPException) as unknown:
        exports.build_export_response("not-a-dataset", "csv")
    assert unknown.value.status_code == 404

    with pytest.raises(HTTPException) as bad_format:
        exports.build_export_response("geri-daily", "pdf")
    assert bad_format.value.status_code == 400


def test_geri_live_and_aligned_market_queries_are_bounded():
    live_query = exports.EXPORT_SPECS["geri-live"].query
    assert "MAX(snapshot_date) - INTERVAL '13 days'" in live_query
    assert "DISTINCT ON (date_trunc('hour', computed_at))" in live_query
    assert "ORDER BY timestamp ASC" in live_query

    brent = exports.EXPORT_SPECS["brent-daily"]
    assert "JOIN intel_indices_daily" in brent.query
    assert brent.params == (exports.GERI_INDEX_ID,)

    vix = exports.EXPORT_SPECS["vix-daily"]
    assert "JOIN intel_indices_daily" in vix.query
    assert "JOIN oil_price_snapshots" in vix.query


def test_country_export_uses_fixed_supported_country_allowlist():
    spec = exports.EXPORT_SPECS["country-storage"]
    assert spec.params == (list(exports.COUNTRY_CODES),)
    assert exports.COUNTRY_CODES == (
        "DE", "FR", "NL", "IT", "AT", "PL", "CZ", "HU", "SK", "BE"
    )
    assert "country_code = ANY(%s)" in spec.query


def test_driver_export_flattens_latest_five_and_preserves_contribution(monkeypatch):
    drivers = [
        {
            "region": "Europe",
            "category": "energy",
            "headline": f"Driver {index}",
            "severity": 5,
            "weighted_score": index / 10,
        }
        for index in range(6)
    ]

    class Cursor:
        def execute(self, query):
            assert "LIMIT 1" in query

        def fetchone(self):
            return {
                "computed_at": datetime(2026, 8, 31, 10, 0),
                "value_change": 2,
                "top_drivers": drivers,
            }

    class Context:
        def __enter__(self):
            return Cursor()

        def __exit__(self, *_args):
            return False

    monkeypatch.setattr(exports, "get_cursor", lambda **_kwargs: Context())
    rows = exports._fetch_driver_rows()
    assert len(rows) == 5
    assert rows[0][1] == "Higher"
    assert rows[3][-1] == pytest.approx(0.3)


def test_driver_export_skips_malformed_items(monkeypatch):
    class Cursor:
        def execute(self, _query):
            pass

        def fetchone(self):
            return {
                "computed_at": datetime(2026, 8, 31, 10, 0),
                "value_change": -1,
                "top_drivers": [
                    "bad",
                    None,
                    {"headline": "Valid 1", "severity": 3},
                    {"headline": "Valid 2", "severity": 2},
                    {"headline": "Valid 3", "severity": 2},
                    {"headline": "Valid 4", "severity": 1},
                    {"headline": "Valid 5", "severity": 1},
                ],
            }

    class Context:
        def __enter__(self):
            return Cursor()

        def __exit__(self, *_args):
            return False

    monkeypatch.setattr(exports, "get_cursor", lambda **_kwargs: Context())
    rows = exports._fetch_driver_rows()
    assert len(rows) == 5
    assert rows[0][1] == "Lower"
    assert rows[0][4] == "Valid 1"
    assert rows[-1][4] == "Valid 5"


def test_csv_and_excel_outputs_have_headers_and_safe_cells(monkeypatch):
    rows = [(datetime(2026, 8, 31, 10, 0), "=unsafe")]
    monkeypatch.setattr(exports, "_fetch_rows", lambda _spec: rows)

    csv_response = exports.build_export_response("geri-live", "csv")
    csv_body = asyncio.run(_response_bytes(csv_response)).decode()
    assert csv_body.startswith("Timestamp,Value")
    assert "'=unsafe" in csv_body
    assert ".csv" in csv_response.headers["content-disposition"]

    xlsx_response = exports.build_export_response("geri-live", "xlsx")
    xlsx_body = asyncio.run(_response_bytes(xlsx_response))
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_body))
    sheet = workbook.active
    assert [sheet.cell(1, column).value for column in (1, 2)] == ["Timestamp", "Value"]
    assert sheet.cell(2, 2).value == "'=unsafe"
    assert sheet.freeze_panes == "A2"


def test_admin_download_route_checks_auth_before_building_export(monkeypatch):
    calls = []

    def reject(token):
        calls.append(("auth", token))
        raise HTTPException(status_code=401, detail="Invalid session")

    monkeypatch.setattr(admin_routes, "verify_admin_token", reject)
    monkeypatch.setattr(
        admin_routes,
        "build_export_response",
        lambda *_args: pytest.fail("export must not run before authorization"),
    )
    with pytest.raises(HTTPException) as error:
        admin_routes.admin_historical_data_download("geri-daily", "csv", None)
    assert error.value.status_code == 401
    assert calls == [("auth", None)]


def test_admin_ui_wires_historical_data_navigation_and_authenticated_downloads():
    html = Path("src/static/admin.html").read_text()
    assert "showSection('historical-data')" in html
    assert 'id="section-historical-data"' in html
    assert "/admin/historical-data/catalog" in html
    assert "/admin/historical-data/download/" in html
    assert "'X-Admin-Token': sessionToken" in html
    assert "setTimeout(() => URL.revokeObjectURL(url), 1000)" in html