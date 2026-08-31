"""Allowlisted historical-data exports for the authenticated admin console."""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Callable, Iterable

from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from src.db.db import get_cursor


@dataclass(frozen=True)
class ExportSpec:
    label: str
    description: str
    group: str
    headers: tuple[str, ...]
    query: str
    params: tuple[Any, ...] = ()


GERI_INDEX_ID = "global:geo_energy_risk"
EERI_INDEX_ID = "europe:eeri"
COUNTRY_CODES = ("DE", "FR", "NL", "IT", "AT", "PL", "CZ", "HU", "SK", "BE")

GERI_DATES_JOIN = """
    JOIN intel_indices_daily geri_dates
      ON geri_dates.date = source.date
     AND geri_dates.index_id = %s
"""

EXPORT_SPECS: dict[str, ExportSpec] = {
    "geri-live": ExportSpec(
        "GERI Live History",
        "Latest 14 calendar days of stored live risk readings.",
        "Risk indices",
        ("Timestamp", "Value"),
        """
        WITH hourly AS (
            SELECT DISTINCT ON (date_trunc('hour', computed_at))
                   computed_at AS timestamp,
                   COALESCE(value_raw, value) AS value
            FROM geri_live_history
            WHERE snapshot_date >= (
                SELECT MAX(snapshot_date) - INTERVAL '13 days' FROM geri_live_history
            )
            ORDER BY date_trunc('hour', computed_at), computed_at DESC
        )
        SELECT timestamp, value
        FROM hourly
        ORDER BY timestamp ASC
        """,
    ),
    "geri-daily": ExportSpec(
        "GERI Daily History",
        "All official daily GERI records with regime.",
        "Risk indices",
        ("Date", "Value", "Regime", "Computed At"),
        """
        SELECT date, value, band AS regime, computed_at
        FROM intel_indices_daily
        WHERE index_id = %s
        ORDER BY date ASC
        """,
        (GERI_INDEX_ID,),
    ),
    "eeri-daily": ExportSpec(
        "EERI Daily History",
        "All official daily European Energy Risk Index records.",
        "Risk indices",
        ("Date", "Value", "Regime", "Computed At"),
        """
        SELECT date, value, band AS regime, computed_at
        FROM reri_indices_daily
        WHERE index_id = %s
        ORDER BY date ASC
        """,
        (EERI_INDEX_ID,),
    ),
    "egsi-m-daily": ExportSpec(
        "EGSI-M Daily History",
        "All European gas-storage macro index records.",
        "Risk indices",
        ("Date", "Value", "Regime", "Computed At"),
        """
        SELECT index_date AS date, index_value AS value, band AS regime, computed_at
        FROM egsi_m_daily
        WHERE region = 'Europe'
        ORDER BY index_date ASC
        """,
    ),
    "egsi-s-daily": ExportSpec(
        "EGSI-S Daily History",
        "All European gas-storage supply index records.",
        "Risk indices",
        ("Date", "Value", "Regime", "Computed At"),
        """
        SELECT index_date AS date, index_value AS value, band AS regime, computed_at
        FROM egsi_s_daily
        WHERE region = 'Europe'
        ORDER BY index_date ASC
        """,
    ),
    "brent-daily": ExportSpec(
        "Brent Historical (Daily)",
        "Daily Brent closes on dates present in GERI Daily.",
        "Market prices",
        ("Date", "Close Value", "Timestamp"),
        f"""
        SELECT source.date, source.brent_price AS close_value, source.created_at AS timestamp
        FROM oil_price_snapshots source
        {GERI_DATES_JOIN}
        WHERE source.brent_price IS NOT NULL
        ORDER BY source.date ASC
        """,
        (GERI_INDEX_ID,),
    ),
    "brent-intraday": ExportSpec(
        "Brent Historical (Intraday)",
        "All stored intraday Brent readings.",
        "Market prices",
        ("Timestamp", "Value"),
        """
        SELECT captured_at AS timestamp, price AS value
        FROM intraday_brent
        ORDER BY captured_at ASC
        """,
    ),
    "vix-daily": ExportSpec(
        "VIX (Daily)",
        "Daily VIX closes on dates shared by GERI Daily and Brent Daily.",
        "Market prices",
        ("Date", "Value", "Timestamp"),
        """
        SELECT source.date, source.vix_close AS value, source.created_at AS timestamp
        FROM vix_snapshots source
        JOIN intel_indices_daily geri_dates
          ON geri_dates.date = source.date AND geri_dates.index_id = %s
        JOIN oil_price_snapshots brent_dates
          ON brent_dates.date = source.date AND brent_dates.brent_price IS NOT NULL
        WHERE source.vix_close IS NOT NULL
        ORDER BY source.date ASC
        """,
        (GERI_INDEX_ID,),
    ),
    "lng-daily": ExportSpec(
        "LNG (Daily)",
        "All stored daily LNG JKM prices.",
        "Market prices",
        ("Date", "Value", "Timestamp"),
        """
        SELECT date, jkm_price AS value, created_at AS timestamp
        FROM lng_price_snapshots
        WHERE jkm_price IS NOT NULL
        ORDER BY date ASC
        """,
    ),
    "ttf-daily": ExportSpec(
        "TTF (Daily)",
        "All stored daily Dutch TTF gas prices.",
        "Market prices",
        ("Date", "Value", "Timestamp"),
        """
        SELECT date, ttf_price AS value, created_at AS timestamp
        FROM ttf_gas_snapshots
        WHERE ttf_price IS NOT NULL
        ORDER BY date ASC
        """,
    ),
    "wti-daily": ExportSpec(
        "WTI (Daily)",
        "All stored daily WTI closes.",
        "Market prices",
        ("Date", "Value", "Timestamp"),
        """
        SELECT date, wti_price AS value, created_at AS timestamp
        FROM oil_price_snapshots
        WHERE wti_price IS NOT NULL
        ORDER BY date ASC
        """,
    ),
    "wti-intraday": ExportSpec(
        "WTI (Intraday)",
        "All stored intraday WTI readings.",
        "Market prices",
        ("Timestamp", "Value"),
        """
        SELECT captured_at AS timestamp, price AS value
        FROM intraday_wti
        ORDER BY captured_at ASC
        """,
    ),
    "natgas-intraday": ExportSpec(
        "Natural Gas (Intraday)",
        "All stored intraday U.S. natural-gas readings.",
        "Market prices",
        ("Timestamp", "Value"),
        """
        SELECT captured_at AS timestamp, price AS value
        FROM intraday_natgas
        ORDER BY captured_at ASC
        """,
    ),
    "eu-storage": ExportSpec(
        "EU Gas Storage (Daily)",
        "All stored EU aggregate gas-storage levels.",
        "Gas storage",
        ("Date", "Value", "Timestamp"),
        """
        SELECT date, eu_storage_percent AS value, created_at AS timestamp
        FROM gas_storage_snapshots
        ORDER BY date ASC
        """,
    ),
    "country-storage": ExportSpec(
        "EU Country Gas Storage",
        "DE, FR, NL, IT, AT, PL, CZ, HU, SK and BE. NL is the source code for the brief's NR entry.",
        "Gas storage",
        ("Country", "Country Code", "Value", "Timestamp"),
        """
        SELECT COALESCE(country_name, country_code) AS country,
               country_code, storage_percent AS value, created_at AS timestamp
        FROM gas_storage_country_snapshots
        WHERE level = 'country' AND country_code = ANY(%s)
        ORDER BY date ASC, country_code ASC
        """,
        (list(COUNTRY_CODES),),
    ),
    "alert-events": ExportSpec(
        "Alert Events",
        "All stored classified alert events supporting the indices.",
        "Risk drivers",
        ("Timestamp", "Alert Type", "Category", "Region", "Severity", "Confidence", "Headline"),
        """
        SELECT created_at AS timestamp, alert_type, category, scope_region AS region,
               severity, confidence, headline
        FROM alert_events
        ORDER BY created_at ASC
        """,
    ),
}

DRIVER_SPEC = ExportSpec(
    "Current GERI Drivers",
    "Latest three to five factors pushing GERI higher or lower; contribution is included when stored.",
    "Risk drivers",
    ("As Of", "GERI Direction", "Region", "Category", "Headline", "Severity", "Contribution"),
    "",
)


def export_catalog() -> list[dict[str, str]]:
    catalog = []
    for key, spec in [*EXPORT_SPECS.items(), ("geri-drivers", DRIVER_SPEC)]:
        catalog.append(
            {
                "key": key,
                "label": spec.label,
                "description": spec.description,
                "group": spec.group,
            }
        )
    return catalog


def _normalise_cell(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _safe_spreadsheet_cell(value: Any) -> Any:
    value = _normalise_cell(value)
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _fetch_rows(spec: ExportSpec) -> list[tuple[Any, ...]]:
    with get_cursor(commit=False) as cur:
        cur.execute(spec.query, spec.params)
        rows = cur.fetchall()
    return [tuple(row[header.lower().replace(" ", "_")] for header in spec.headers) for row in rows]


def _fetch_driver_rows() -> list[tuple[Any, ...]]:
    with get_cursor(commit=False) as cur:
        cur.execute(
            """
            SELECT computed_at, value_change, top_drivers
            FROM geri_live_history
            WHERE jsonb_typeof(top_drivers) = 'array'
              AND jsonb_array_length(top_drivers) > 0
            ORDER BY computed_at DESC
            LIMIT 1
            """
        )
        row = cur.fetchone()
    if not row:
        return []
    change = row.get("value_change")
    direction = "Higher" if change and change > 0 else ("Lower" if change and change < 0 else "Flat")
    drivers = row.get("top_drivers") or []
    if isinstance(drivers, str):
        try:
            drivers = json.loads(drivers)
        except (TypeError, ValueError):
            return []
    if not isinstance(drivers, list):
        return []
    result = []
    for driver in drivers:
        if not isinstance(driver, dict):
            continue
        result.append(
            (
                row.get("computed_at"),
                direction,
                driver.get("region"),
                driver.get("category"),
                driver.get("headline"),
                driver.get("severity"),
                driver.get("contribution", driver.get("weighted_score")),
            )
        )
        if len(result) == 5:
            break
    return result


def _csv_response(spec: ExportSpec, rows: Iterable[tuple[Any, ...]], filename: str):
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(spec.headers)
    for row in rows:
        writer.writerow([_safe_spreadsheet_cell(value) for value in row])
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


def _xlsx_response(spec: ExportSpec, rows: Iterable[tuple[Any, ...]], filename: str):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = spec.label[:31]
    fill = PatternFill("solid", fgColor="1E293B")
    font = Font(bold=True, color="F1F5F9")
    for column, header in enumerate(spec.headers, 1):
        cell = sheet.cell(1, column, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    for row_number, row in enumerate(rows, 2):
        for column, value in enumerate(row, 1):
            sheet.cell(row_number, column, _safe_spreadsheet_cell(value))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, header in enumerate(spec.headers, 1):
        sheet.column_dimensions[get_column_letter(column)].width = min(max(len(header) + 4, 14), 48)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def build_export_response(dataset: str, file_format: str):
    if file_format not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="Format must be csv or xlsx")
    if dataset == "geri-drivers":
        spec = DRIVER_SPEC
        rows = _fetch_driver_rows()
    else:
        spec = EXPORT_SPECS.get(dataset)
        if not spec:
            raise HTTPException(status_code=404, detail="Unknown historical dataset")
        rows = _fetch_rows(spec)
    filename = f"{dataset.replace('-', '_')}_{date.today().isoformat()}"
    renderer: Callable[..., Any] = _csv_response if file_format == "csv" else _xlsx_response
    return renderer(spec, rows, filename)